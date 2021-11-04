from openpyxl import load_workbook
filename = str(input("type a name of an existing file: "))
wb1 = load_workbook(filename+'.xlsx')
ws1 = wb1.active
countrow = 0
for x in ws1.iter_rows(): 
    match_num = 0
    for y in range(1,len(x),4):
        name = x[0].value
        reference = x[y].value
        result = x[y+2].value         
        if reference == result:
            match_num = match_num + 1
    lastcol = len(x)+1
    countrow = countrow+1
    c = ws1.cell(row=countrow, column=lastcol, value=match_num)
print("Ta dam!")
wb1.save(filename+'.xlsx')
idle = input("")
